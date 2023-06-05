import shutil
import time
import openpyxl
import os

# 원본 파일 경로와 복사될 파일 경로 지정
original_file_path = "C:/inersection/4intersection/4intersection/4intersection/4intersection/intersection.xlsx"
copied_file_path ="C:/inersection/4intersection/4intersection/4intersection/4intersection/intersection_copy.xlsx"

def check_file_exists(file_path):
    return os.path.exists(file_path)

def check_sheet_exists(workbook, sheet_name):
    return sheet_name in workbook.sheetnames

while True:
    try:
        shutil.copy(original_file_path, copied_file_path)
        print('복사 완료')
    except:
        print('복사 실패')
    time.sleep(1)

    try:
        if not check_file_exists(copied_file_path):
            print('복사된 파일이 존재하지 않습니다.')
            time.sleep(1)
            continue

        wb = openpyxl.load_workbook(copied_file_path)
        if not check_sheet_exists(wb, 'car_coordinateX'):
            print('car_coordinateX 시트가 존재하지 않습니다.')
            time.sleep(1)
            continue

        ws = wb['car_coordinateX']

        # 시작 열의 인덱스, 시작 행의 인덱스, 결과를 저장할 리스트
        start_col = 2
        start_row = 2
        result_list = []

        # 일정 시간 간격으로 행 인덱스를 이동하면서 값을 출력
        while True:
            # 지정한 행의 값을 가져와서 리스트에 추가
            row_values = []
            for idx, cell in enumerate(ws[start_row]):
                if cell.value is None:
                    row_values.append(None)
                else:
                    row_values.append(cell.value)

            # 결과 리스트에 추가
            result_list.append(row_values)

            # 결과 출력
            for idx, val in enumerate(row_values):
                if val is not None:
                    if idx == 0:
                        print(f'{idx} : {val}초')
                    else:
                        print(f'{idx}번 차량 : {val}')

            # 일정 시간(3초) 대기 후 행 인덱스 조절
            time.sleep(3)
            print('------------------')
            start_row += 2  # 2씩 증가

            # 원본 파일과 복사 파일이 다르면 복사한 파일을 닫고 다시 복사합니다.
            if original_file_path != copied_file_path:
                wb.close()
                while True:
                    try:
                        shutil.copy(original_file_path, copied_file_path)
                        print('재복사 완료')
                        break  # 복사에 성공하면 while문 탈출
                    except:
                        print('파일이 잠겨 있습니다. 잠시 대기합니다.')
                        time.sleep(1)
                wb = openpyxl.load_workbook(copied_file_path)
                ws = wb['car_coordinateX']

    except:
        print('데이터 추출 실패')

    time.sleep(2)  # 2초간 대기
